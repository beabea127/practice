import csv
import os
import urllib.request

# ===== 設定 =====
CSV_PATH = r"C:\Users\YourName\Documents\pdf_links.csv"
OUTPUT_DIR = r"C:\Users\YourName\Documents\downloaded_pdfs"


def download_pdf(url, save_path):
    """PDFをダウンロードし、成功ならTrue、失敗ならFalseを返す"""
    try:
        request = urllib.request.Request(
            url,
            headers={"User-Agent": "Mozilla/5.0"}  # ブロック防止
        )
        with urllib.request.urlopen(request, timeout=10) as response:
            if response.status != 200:
                return False

            with open(save_path, "wb") as f:
                f.write(response.read())

        return True

    except Exception:
        return False


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    with open(CSV_PATH, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)

        for row_idx, row in enumerate(reader, start=1):

            # C列 = 3番目
            if len(row) < 3:
                continue

            url = row[2].strip()
            if not url:
                continue

            # PDFで終わらないURLはスキップ（必要なら削除OK）
            if not url.lower().endswith(".pdf"):
                print(f"行{row_idx}: PDFリンクではない → スキップ: {url}")
                continue

            # 保存先
            save_path = os.path.join(OUTPUT_DIR, f"row_{row_idx}.pdf")

            print(f"行{row_idx}: ダウンロード中 → {url}")

            # ここで失敗時にスキップ
            if download_pdf(url, save_path):
                print(f"  ✔ 保存成功: {save_path}")
            else:
                print(f"  ✖ ダウンロード失敗（URLが無効？）→ スキップ: {url}")

    print("\nすべての処理が完了しました。")


if __name__ == "__main__":
    main()



####


from openpyxl import load_workbook

# === 設定 ===
excel_path = r"C:\Users\Suguru Abe\Desktop\your_file.xlsx"  # Excelファイルのパス
sheet_name = None   # None = 最初のシートを使う
col = "C"           # C列のURLを取り出す
output_txt = r"C:\Users\Suguru Abe\Desktop\extracted_urls.txt"

# === Excelを読み込み ===
wb = load_workbook(excel_path, data_only=True)
ws = wb[sheet_name] if sheet_name else wb.active

urls = []

# === C列を上から走査 ===
for row in range(1, ws.max_row + 1):
    cell = ws[f"{col}{row}"]
    if cell.hyperlink:
        urls.append(cell.hyperlink.target)  # URL（リンク先）を取得

# === 結果を保存 ===
with open(output_txt, "w", encoding="utf-8") as f:
    for u in urls:
        f.write(u + "\n")

print("抽出完了！ 件数:", len(urls))
print("保存先:", output_txt)


####

import os
from pathlib import Path
from urllib.parse import urlparse, unquote
import requests

# ==== 設定 ====
URL_LIST_FILE = r"C:\Users\Suguru Abe\Desktop\extracted_urls.txt"  # URLリスト
OUTPUT_DIR    = r"C:\Users\Suguru Abe\Desktop\downloaded_pdfs"     # 保存先フォルダ
TIMEOUT       = 60  # 秒：レスポンス待ち時間
# ==============

Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)

def make_safe_filename(name: str) -> str:
    """Windowsで使えない文字を置き換える"""
    bad_chars = '\\/:*?"<>|'
    for ch in bad_chars:
        name = name.replace(ch, "_")
    return name

def filename_from_response(url: str, resp: requests.Response, index: int) -> str:
    """レスポンスヘッダ or URL からPDFファイル名を決める"""
    # 1. Content-Disposition から取得を試みる
    cd = resp.headers.get("Content-Disposition", "")
    if "filename=" in cd:
        # filename="xxx.pdf" 形式をざっくり抜き出し
        fname = cd.split("filename=")[-1].strip().strip('";')
        fname = unquote(fname)
    else:
        # 2. URLのパスから取得
        path = urlparse(url).path
        fname = Path(path).name
        fname = unquote(fname)

    # 3. 何も取れなければ index を使う
    if not fname:
        fname = f"file_{index:04d}.pdf"

    # 4. 拡張子が付いていなければ .pdf を付ける
    if not fname.lower().endswith(".pdf"):
        fname += ".pdf"

    return make_safe_filename(fname)

# ==== URLリストを読み込む ====
with open(URL_LIST_FILE, "r", encoding="utf-8") as f:
    urls = [line.strip() for line in f if line.strip()]

print(f"URL件数: {len(urls)}")

# ==== 1件ずつPDFをダウンロード ====
for i, url in enumerate(urls, start=1):
    print(f"[{i}/{len(urls)}] Downloading: {url}")

    try:
        resp = requests.get(url, timeout=TIMEOUT)
        resp.raise_for_status()
    except Exception as e:
        print(f"  ❌ ダウンロード失敗: {e}")
        continue

    # 一応Content-TypeがPDFかチェック（違っても保存したければコメントアウト）
    ctype = resp.headers.get("Content-Type", "")
    if "pdf" not in ctype.lower():
        print(f"  ⚠ Content-TypeがPDFではありません ({ctype}) → スキップ")
        continue

    filename = filename_from_response(url, resp, i)
    out_path = Path(OUTPUT_DIR) / filename

    # 同名ファイルがあれば連番を付けて回避
    cnt = 1
    base_stem = out_path.stem
    suffix = out_path.suffix
    while out_path.exists():
        out_path = Path(OUTPUT_DIR) / f"{base_stem}_{cnt}{suffix}"
        cnt += 1

    try:
        with open(out_path, "wb") as f:
            f.write(resp.content)
        print(f"  ✅ 保存: {out_path}")
    except Exception as e:
        print(f"  ❌ 保存中にエラー: {e}")

print("完了しました。")

#####
pip install openpyxl
pip install requests

pip install --no-index --find-links=. et_xmlfile-1.1.0-py3-none-any.whl
pip install --no-index --find-links=. openpyxl-3.1.2-py2.py3-none-any.whl
pip install --no-index --find-links=. requests-2.32.3-py3-none-any.whl
